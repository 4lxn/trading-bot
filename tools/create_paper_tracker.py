#!/usr/bin/env python3
"""Generates paper_trading_8b.xlsx — the manual paper-trading log for
strategy 8b, with automatic equity, P&L and drawdown formulas.

Run from the repo root:  python3 tools/create_paper_tracker.py
"""

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "paper_trading_8b.xlsx")

N_ROWS = 500
FIRST = 4  # first data row
HEADERS = [
    ("A", "Date", 12),
    ("B", "Signal (LONG/OUT)", 16),
    ("C", "Action (BUY/SELL/-)", 17),
    ("D", "Price (USDT)", 14),
    ("E", "BTC after", 12),
    ("F", "USDT after", 12),
    ("G", "Equity (USDT)", 14),
    ("H", "P&L %", 10),
    ("I", "Peak equity", 12),
    ("J", "Drawdown %", 12),
]

wb = Workbook()

# --- Log sheet --------------------------------------------------------------
ws = wb.active
ws.title = "Log"
ws["A1"] = "Initial capital (USDT):"
ws["A1"].font = Font(bold=True)
ws["B1"] = 1000
ws["B1"].fill = PatternFill("solid", fgColor="FFF2CC")

header_fill = PatternFill("solid", fgColor="D9E1F2")
for col, title, width in HEADERS:
    cell = ws[f"{col}3"]
    cell.value = title
    cell.font = Font(bold=True)
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions[col].width = width

for r in range(FIRST, FIRST + N_ROWS):
    # Equity = BTC after * price + USDT after
    ws[f"G{r}"] = f'=IF(D{r}="","",E{r}*D{r}+F{r})'
    # Cumulative P&L vs initial capital
    ws[f"H{r}"] = f'=IF(G{r}="","",G{r}/$B$1-1)'
    ws[f"H{r}"].number_format = "0.0%"
    # Running peak equity and drawdown from peak
    ws[f"I{r}"] = f'=IF(G{r}="","",MAX(G${FIRST}:G{r}))'
    ws[f"J{r}"] = f'=IF(G{r}="","",G{r}/I{r}-1)'
    ws[f"J{r}"].number_format = "0.0%"

ws.freeze_panes = "A4"

# --- Instructions sheet -------------------------------------------------------
info = wb.create_sheet("Instructions")
info.column_dimensions["A"].width = 100
lines = [
    "Paper trading log — strategy 8b (long while close > EMA200 and RSI(14) > 55, daily, long-only)",
    "",
    "1. Set your starting capital in Log!B1.",
    "2. Add one row per day you check the signal (or at least per signal change):",
    "   - Date: the daily candle date (UTC close).",
    "   - Signal: LONG or OUT, as reported by the bot / backtest.py.",
    "   - Action: BUY when the signal turns LONG, SELL when it turns OUT, '-' otherwise.",
    "   - Price: BTC price used for the (paper) fill.",
    "   - BTC after / USDT after: your paper balances after the action.",
    "3. Equity, P&L %, peak and drawdown compute automatically.",
    "",
    "Purpose: build weeks of honest, side-by-side evidence (paper vs bot logs)",
    "before risking real money. Expect drawdowns of -30% to -50% at some point.",
    "This is educational material, not financial advice.",
]
for i, line in enumerate(lines, start=1):
    info[f"A{i}"] = line
info["A1"].font = Font(bold=True)

wb.save(os.path.normpath(OUT))
print(f"Wrote {os.path.normpath(OUT)}")
